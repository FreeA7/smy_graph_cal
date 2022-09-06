# -*- coding: UTF-8 -*-
import os
import openpyxl

# 输入文件名
INPUT_FILE_NAME = "Centrality Network1.xlsx"
# 导出的文件夹名称
OUTPUT_FOLDER_NAME = "output"


def main():
    """
    程序入口
    :return:
    """
    # 读取excel数据
    title, datas = read_excel(INPUT_FILE_NAME)
    # 没有文件夹的创建文件夹
    if not os.path.exists(OUTPUT_FOLDER_NAME):
        os.makedirs(OUTPUT_FOLDER_NAME)
    # 逐个写入文件
    for ygrp in datas:
        file_path = os.path.join(OUTPUT_FOLDER_NAME, ygrp + ".xlsx")
        save_data(file_path, title, datas[ygrp])


def read_excel(file_name):
    """
    读取Excel，返回表头和以第二列为key，相同第二列的所有行的列表为value的字典
    :param file_name:
    :return:
    """
    wb = openpyxl.load_workbook(file_name, read_only=True)
    sheet = wb.active
    # 读取标题
    title = []
    # 数据
    datas = {}
    for i, row in enumerate(sheet.rows):
        # 第一行为表头
        if i == 0:
            title = [col.value for col in row]
        # 其余行
        else:
            # key
            ygrp = row[1].value.strip()
            if ygrp not in datas:
                datas[ygrp] = []
            # value
            datas[ygrp].append([col.value for col in row])
    wb.close()
    return title, datas


def save_data(file_name, title, rows):
    """
    将结果数据写入file_name的Excel文件
    :param file_name:
    :param title:
    :param rows:
    :return:
    """
    if not file_name or not title or not rows:
        return
    print(file_name)
    wb = openpyxl.Workbook()
    sheet = wb.active
    # 写标题
    for i in range(len(title)):
        sheet.cell(row=1, column=i+1, value=title[i])
    # 写内容
    for i, row in enumerate(rows):
        print(row)
        for j, col in enumerate(row):
            sheet.cell(row=i+2, column=j+1, value=col)
    wb.save(filename=file_name)
    wb.close()


if __name__ == "__main__":
    main()
